#This is the support library for TDA Online. It contains various functions for processing file analysis requests as passed to it by the Flask app, unified under the 'analyze' function.
#The accompanying Flask app sends a file in either a TextIO or ByteIO wrapper with information on file type as determined by magic. Dimensions of the data are automatically determined and data is extracted and organized into a NumPy array.
#Analysis is through ripser, taking the maximum feature dimension as the minimum of the dimension of the data and the maxdim value passed by the Flask app. Betti numbers and the Euler characteristics are calculated at every change, and statistical analysis is performed.
#Stats are provided for each dimension of nonempty results in the following presentation: [number of features, min length, max length, mean length, median length]; max, mean, and median calculations ignore the 0-dimensional feature of infinite length. The lists are collected in the superlist 'stats' in increasing dimension order.
#Ratios are calculated if the resultant features span at least 2 dimensions. Ratios are ordered in 'lexicographical order' (0:1, 0:2, 1:2) with ratios involving any empty set of features omitted. The resulting list is appended as the last item to 'stats.'
#Support is currently for CSV files with the capability to accept any delimiter and any line terminator and for Excel files. There are plans to add support for other files including PDB (Protein Data Bank) files.

#import packages and set Matplotlib to output to buffer
from io import BytesIO
from copy import copy
from statistics import mean, median
import numpy
from ripser import ripser
import matplotlib
import matplotlib.pyplot
from csv import reader
from openpyxl import load_workbook
matplotlib.use('Agg')

#This is he function that accepts all input from the Flask app. Provided the file data and the file type, it first directs files for data extraction appropriately. In the case an unsupported file is submitted, automatically returns.
#Next, the function determines the dimension of the data and picks the minimum between it and the max dimension specified. This is done to avoid attempting analysis for nonexistent dimensions, while still respecting that only the requested analysis is performed.
#The simplicial complex is computed, and dimensions with features are determined.
#Next, it calls functions to get the betti numbers, Euler characteristics, and statistics.
#The next sequence of steps involving plot creation are performed once for the persistence diagram and once for the persistence barcode:
#A variable is prepared as a BytesIO object to accept a byte stream. This is where the plot will be written so that it can be served to the client without being first written to any non-volatile server storage.
#The data is plotted and the final plot written to the BytesIO variable, then the plot is cleared.
#The data is returned to the Flask app as a list in the following way: [persistence diagram, persistence barcode, betti numbers and Euler characteristics, statistics, simplicial complex, list of dimensions with features]
def analyze(file, type, maxdim=3, coeff=7, delimiter=',', lineterminator='\n', igLabels=False, igEnum=False, maxSize=200):
  try:
    if type == 'csv': data = extractCSV(file, delimiter, lineterminator)
    if type == 'xls': data = extractExcel(file)
  except: return ['File format error.']
  if igLabels: data = numpy.delete(data, (0), axis=0)
  if len(data) > maxSize: return ['Data too large! Please submit a data set with ' + str(maxSize) + ' points or fewer (your file contains ' + str(len(data)) + ' points).']
  if igEnum: data = numpy.delete(data, (0), axis=1)
  maxd = len(data[0])
  if maxdim < maxd: maxd = maxdim
  dgms = ripser(data, coeff=coeff, maxdim=maxd)['dgms']
  nontrivd = []
  for i in range(1, len(dgms)):
    if dgms[i].size > 0: nontrivd.append(i)
  totald = copy(nontrivd)
  totald.append(0)
  totald.sort()
  betti = getBetti(dgms, nontrivd, totald)
  stats = getStats(dgms, nontrivd, totald)
  buf = BytesIO()
  for i in dgms:
    matplotlib.pyplot.scatter(i[:,0], i[:,1])
  m = max(dgms[0][0:-2][:,1])
  for i in nontrivd:
    m = max(m, max(dgms[i][:,1]))
  matplotlib.pyplot.legend(totald, loc='lower right')
  matplotlib.pyplot.plot((0, m), (0, m), color ='black', linestyle='dashed')
  matplotlib.pyplot.savefig(buf, bbox_inches='tight', format='svg')
  matplotlib.pyplot.clf()
  colours=['black', 'blue', 'red', 'purple', 'green']
  bar = BytesIO()
  step = 0
  for i in totald:
    for j in range(len(dgms[i])):
      matplotlib.pyplot.plot((dgms[i][-j-1][0], dgms[i][-j-1][1]), (step, step), color=colours[i], linestyle='solid')
      step += 1
  legend=[]
  for i in range(len(totald)):
    legend.append(matplotlib.patches.Patch(color=colours[totald[i]], label=totald[i]))
  matplotlib.pyplot.legend(handles=legend)
  matplotlib.pyplot.savefig(bar, bbox_inches='tight', format='svg')
  matplotlib.pyplot.clf()
  return [buf, bar, betti, stats, dgms, totald]

#The getStats funtion accepts the simplicial complex and produces and returns the stats and ratios described in the overview at the top.
def getStats(dgms, nontrivd, totald):
  stats = [[len(dgms[0])]]
  lifetimes = []
  for i in range(len(dgms[0])-1):
    lifetimes.append(dgms[0][i][1])
  for k in [min(lifetimes), max(lifetimes), mean(lifetimes), median(lifetimes)]:
    stats[0].append(k)
  for j in nontrivd:
    lifetimes = []
    for i in range(len(dgms[j])):
      lifetimes.append(dgms[j][i][1]-dgms[j][i][0])
    stats.append([len(dgms[j])])
    for k in [min(lifetimes), max(lifetimes), mean(lifetimes), median(lifetimes)]:
      stats[j].append(k)
  ratios = []
  for i in totald:
    for j in nontrivd:
      if j <= i: continue
      ratios.append(stats[i][0]/stats[j][0])
      ratios.append(stats[i][3]/stats[j][3])
  stats.append(ratios)
  return stats

#The getBetti function accepts the simplicial complex and calculates the betti numbers and Euler characteristic every time they change.
def getBetti(dgms, nontrivd, totald):
  tchanges = []
  for i in range(len(dgms[0])-1):
    tchanges.append(dgms[0][i][0])
    tchanges.append(dgms[0][i][1])
  for i in nontrivd:
    for j in range(len(dgms[i])):
      tchanges.append(dgms[i][j][0])
      tchanges.append(dgms[i][j][1])
  tchanges = sorted(set(tchanges))
  betti = [tchanges]
  for i in totald:
    count = []
    for j in tchanges:
      total = 0
      for k in range(len(dgms[i])):
        if dgms[i][k][0] <= j: total += 1
        if dgms[i][k][1] <= j: total -= 1
      count.append(total)
    betti.append(count)
  ec= []
  for i in range(len(tchanges)):
    total = 0
    for j in range(len(totald)):
      if totald[j]%2 == 0 : total += betti[j+1][i]
      else: total -= betti[j+1][i]
    ec.append(total)
  betti.append(ec)
  return betti

#The extractExcel function extracts and prepares data from Excel files.
#Accepts data held in a BufferedReader(BytesIO) wrapper stack.
def extractExcel(file):
  wb = load_workbook(file)
  sheet = wb[wb.sheetnames[0]]
  data = numpy.array(list(sheet.values))
  return data

#The extractCVS function extracts and prepares data from CSV files.
#Accepts data held in a TextIO(BufferedReader(BytesIO)) wrapper stack, the delimiter, and the line terminator.
def extractCSV(file, delimiter, lineterminator):
  table = list(reader(file, delimiter=delimiter, lineterminator=lineterminator))
  l = len(table)
  d = len(table[0])
  data = []
  for i in range(l):
    coordinates = []
    for j in range(d):
      coordinates.append(table[i][j])
    data.append(coordinates)
  data = numpy.array(data)
  return data
